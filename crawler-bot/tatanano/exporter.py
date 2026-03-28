"""
ScrapeNano — Excel Exporter
Exports leads from SQLite to a formatted .xlsx file.
"""
import os
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from . import config, db

# ── Colour palette ─────────────────────────────────────────────────────────────
C_HEADER_BG  = "1A1A2E"   # dark navy
C_HEADER_FG  = "FFFFFF"   # white
C_ROW_ALT    = "F5F0FA"   # light lavender for alternating rows
C_ACCENT     = "6B2FA0"   # purple accent
C_BORDER     = "D0C8E0"   # soft border
C_TITLE_BG   = "2D1042"   # dark purple for title row

# ── Column definitions ─────────────────────────────────────────────────────────
COLUMNS = [
    ("name",        "Name",         22),
    ("email",       "Email",        28),
    ("phone",       "Phone",        18),
    ("company",     "Company",      24),
    ("title",       "Job Title",    22),
    ("location",    "Location",     20),
    ("source_url",  "Source URL",   35),
]


def export_job(job_id: int) -> str:
    """
    Export all leads for a job to Excel.
    Returns the file path of the created .xlsx file.
    """
    job = db.get_job(job_id)
    if not job:
        raise ValueError(f"Job {job_id} not found")

    leads = db.get_leads(job_id)
    job_name_safe = _safe_filename(job["name"])
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"scrapenano_{job_name_safe}_{ts}.xlsx"
    filepath = str(Path(config.EXPORT_DIR) / filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"

    # ── Title row ────────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(len(COLUMNS))}1")
    title_cell = ws["A1"]
    title_cell.value = f"ScrapeNano Export — {job['name']}"
    title_cell.font = Font(name="Calibri", size=14, bold=True, color=C_HEADER_FG)
    title_cell.fill = PatternFill("solid", fgColor=C_TITLE_BG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # ── Meta row ─────────────────────────────────────────────────────────────
    ws.merge_cells(f"A2:{get_column_letter(len(COLUMNS))}2")
    meta_cell = ws["A2"]
    meta_cell.value = (
        f"Job: {job['description'][:80]}{'...' if len(job['description']) > 80 else ''}  |  "
        f"Leads: {len(leads)}  |  Exported: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    )
    meta_cell.font = Font(name="Calibri", size=9, italic=True, color="AAAAAA")
    meta_cell.fill = PatternFill("solid", fgColor="0F0F1E")
    meta_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 18

    # ── Header row ────────────────────────────────────────────────────────────
    header_row = 3
    thin = Side(style="thin", color=C_BORDER)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_i, (_, label, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_i, value=label)
        cell.font = Font(name="Calibri", size=10, bold=True, color=C_HEADER_FG)
        cell.fill = PatternFill("solid", fgColor=C_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(col_i)].width = width
    ws.row_dimensions[header_row].height = 22

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_i, lead in enumerate(leads, start=header_row + 1):
        alt = (row_i % 2 == 0)
        row_fill = PatternFill("solid", fgColor=C_ROW_ALT) if alt else None

        for col_i, (field, _, _) in enumerate(COLUMNS, start=1):
            value = lead.get(field) or lead.get("data", {}).get(field, "")
            cell = ws.cell(row=row_i, column=col_i, value=value or "")
            cell.font = Font(name="Calibri", size=9)
            cell.alignment = Alignment(vertical="center", wrap_text=(field == "source_url"))
            cell.border = border
            if row_fill:
                cell.fill = row_fill

            # Hyperlink for emails
            if field == "email" and value:
                cell.hyperlink = f"mailto:{value}"
                cell.font = Font(name="Calibri", size=9, color=C_ACCENT, underline="single")
            # Hyperlink for URLs
            if field == "source_url" and value:
                cell.hyperlink = value
                cell.font = Font(name="Calibri", size=9, color="0070C0", underline="single")

        ws.row_dimensions[row_i].height = 16

    # ── Freeze panes ──────────────────────────────────────────────────────────
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # ── Auto-filter ───────────────────────────────────────────────────────────
    ws.auto_filter.ref = (
        f"A{header_row}:{get_column_letter(len(COLUMNS))}{header_row + len(leads)}"
    )

    # ── Stats sheet ───────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Stats")
    _write_stats_sheet(ws2, job, leads)

    wb.save(filepath)
    return filepath


def _write_stats_sheet(ws, job: dict, leads: list[dict]):
    """Simple stats summary on second sheet."""
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 40

    rows = [
        ("Job Name",     job["name"]),
        ("Description",  job["description"]),
        ("Created At",   job["created_at"]),
        ("Total Leads",  len(leads)),
        ("With Email",   sum(1 for l in leads if l.get("email"))),
        ("With Phone",   sum(1 for l in leads if l.get("phone"))),
        ("With Company", sum(1 for l in leads if l.get("company"))),
        ("Unique Domains", _count_domains(leads)),
    ]

    for r, (k, v) in enumerate(rows, start=1):
        ws.cell(row=r, column=1, value=k).font = Font(bold=True, size=10)
        ws.cell(row=r, column=2, value=str(v)).font = Font(size=10)


def _count_domains(leads: list[dict]) -> int:
    domains = set()
    for l in leads:
        url = l.get("source_url") or ""
        if "://" in url:
            domains.add(url.split("/")[2])
    return len(domains)


def _safe_filename(name: str) -> str:
    import re
    return re.sub(r'[^\w\-]', '_', name)[:30]
